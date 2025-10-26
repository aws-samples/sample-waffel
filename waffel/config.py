#!/usr/bin/env python3

from openpyxl.styles import Font, PatternFill, Alignment

class WAFFELConfig:
    """Configuration settings for WAFFEL"""
    
    # Pillar name mappings
    PILLAR_NAMES = {
        'operationalExcellence': 'Operational Excellence',
        'security': 'Security',
        'reliability': 'Reliability', 
        'performance': 'Performance Efficiency',
        'costOptimization': 'Cost Optimization',
        'sustainability': 'Sustainability'
    }
    
    PILLAR_ABBREVIATIONS = {
        'Operational Excellence': 'OPS',
        'Security': 'SEC',
        'Reliability': 'REL',
        'Performance Efficiency': 'PERF',
        'Cost Optimization': 'COST',
        'Sustainability': 'SUS'
    }
    
    # Excel styles
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    QUESTION_FONT = Font(bold=True)
    QUESTION_FILL = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    # Risk colors
    RISK_COLORS = {
        'High Risk': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'Medium Risk': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        'Low Risk': PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")
    }
    
    # Status colors
    STATUS_COLORS = {
        '✅ Selected': PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid"),
        '⚠️ Not Selected': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        'Not Applicable': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    }
    
    # Summary risk colors
    HIGH_RISK_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    MEDIUM_RISK_FILL = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
    
    # Column widths
    COLUMN_WIDTHS = {
        'workload_props': {'A': 20, 'B': 80},
        'summary': {'A': 25, 'others': 12},
        'improvement': {'A': 20, 'B': 15, 'C': 50, 'D': 15, 'E': 60, 'F': 40},
        'pillar': {'A': 12, 'B': 40, 'C': 20, 'D': 12, 'E': 30, 'F': 40}
    }
    
    # Sheet names and headers
    SHEET_NAMES = {
        'properties': 'Workload properties',
        'summary': 'Summary',
        'improvement': 'Improvement plan'
    }
    
    HEADERS = {
        'properties': ['Property', 'Value'],
        'summary': ['Pillar', 'Questions Answered', 'High Risk', 'Medium Risk', 'Low Risk', 'With Notes', 'Selected', 'Not Selected', 'N/A'],
        'improvement': ['Pillar', 'Question ID', 'Question', 'Risk Level', 'Improvement Item', 'URL'],
        'pillar': ['Question ID', 'Question', 'Choice/Details', 'Risk Level', 'Notes', 'Improvement Plan']
    }
    
    # Property display order
    PROPERTY_ORDER = ['Workload name', 'ARN', 'Description', 'Review owner', 'Environment', 'AWS Regions', 'Account IDs', 'Date']
    
    # API risk level mapping
    API_RISK_MAPPING = {
        'HIGH': 'High Risk',
        'MEDIUM': 'Medium Risk', 
        'LOW': 'Low Risk',
        'UNANSWERED': 'Not Assessed',
        'NOT_APPLICABLE': 'Not Applicable'
    }
    
    # Pillar colors for PowerPoint (RGB values) - Pastel colors
    PILLAR_COLORS = {
        'Operational Excellence': (255, 204, 153),   # Light Orange
        'Security': (255, 153, 153),                 # Light Red
        'Reliability': (153, 255, 153),              # Light Green
        'Performance Efficiency': (153, 204, 255),  # Light Blue
        'Cost Optimization': (255, 255, 153),       # Light Yellow
        'Sustainability': (204, 153, 255)           # Light Purple
    }
    
    # High risk border color
    HIGH_RISK_BORDER_COLOR = (255, 0, 0)  # Red border for high risk
