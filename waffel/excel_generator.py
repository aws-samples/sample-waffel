#!/usr/bin/env python3

from openpyxl import Workbook
from openpyxl.styles import Alignment
from .config import WAFFELConfig

class ExcelGenerator:
    """Generates Excel files from Well-Architected data"""
    
    def __init__(self):
        self.config = WAFFELConfig()
    
    def generate(self, pillars, workload_props, output_path):
        """Generate Excel file from standardized data"""
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        self._create_workload_properties_sheet(wb, workload_props)
        self._create_summary_sheet(wb, pillars)
        self._create_improvement_plan_sheet(wb, pillars)
        self._create_pillar_sheets(wb, pillars)
        
        wb.save(output_path)
    
    def _create_workload_properties_sheet(self, wb, workload_props):
        """Create workload properties sheet"""
        ws = wb.create_sheet(self.config.SHEET_NAMES['properties'])
        
        # Headers
        ws.append(self.config.HEADERS['properties'])
        
        # Format headers
        for cell in ws[1]:
            cell.font = self.config.HEADER_FONT
            cell.fill = self.config.HEADER_FILL
            cell.alignment = self.config.CENTER_ALIGNMENT
        
        # Add properties in order
        for prop in self.config.PROPERTY_ORDER:
            if prop in workload_props:
                ws.append([prop, workload_props[prop]])
        
        # Add remaining properties
        for prop, value in workload_props.items():
            if prop not in self.config.PROPERTY_ORDER:
                ws.append([prop, value])
        
        # Set column widths and formatting
        ws.column_dimensions['A'].width = self.config.COLUMN_WIDTHS['workload_props']['A']
        ws.column_dimensions['B'].width = self.config.COLUMN_WIDTHS['workload_props']['B']
        
        # Set row height and alignment
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 15
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, vertical="center")
    
    def _create_summary_sheet(self, wb, pillars):
        """Create summary sheet"""
        ws = wb.create_sheet(self.config.SHEET_NAMES['summary'])
        
        # Headers
        ws.append(self.config.HEADERS['summary'])
        
        # Format headers
        for cell in ws[1]:
            cell.font = self.config.HEADER_FONT
            cell.fill = self.config.HEADER_FILL
            cell.alignment = self.config.CENTER_ALIGNMENT
        
        # Add summary data
        for pillar_name, questions in pillars.items():
            if questions:
                total_questions = len(questions)
                answered_questions = len([q for q in questions if q['stats']['selected'] > 0])
                
                high_risk = len([q for q in questions if q['risk_level'] == 'High Risk'])
                medium_risk = len([q for q in questions if q['risk_level'] == 'Medium Risk'])
                low_risk = len([q for q in questions if q['risk_level'] == 'Low Risk'])
                with_notes = len([q for q in questions if q['notes']])
                
                total_selected = sum(q['stats']['selected'] for q in questions)
                total_not_selected = sum(q['stats']['not_selected'] for q in questions)
                total_na = sum(q['stats']['na'] for q in questions)
                
                ws.append([
                    pillar_name,
                    f"{answered_questions}/{total_questions}",
                    high_risk,
                    medium_risk,
                    low_risk,
                    with_notes,
                    total_selected,
                    total_not_selected,
                    total_na
                ])
        
        # Format data and apply risk colors
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = self.config.CENTER_ALIGNMENT
            
            if len(row) >= 5:
                high_risk_cell = row[2]
                medium_risk_cell = row[3]
                
                if high_risk_cell.value and high_risk_cell.value > 0:
                    high_risk_cell.fill = self.config.HIGH_RISK_FILL
                if medium_risk_cell.value and medium_risk_cell.value > 0:
                    medium_risk_cell.fill = self.config.MEDIUM_RISK_FILL
        
        # Set column widths
        ws.column_dimensions['A'].width = self.config.COLUMN_WIDTHS['summary']['A']
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = self.config.COLUMN_WIDTHS['summary']['others']
    
    def _create_improvement_plan_sheet(self, wb, pillars):
        """Create improvement plan sheet"""
        ws = wb.create_sheet(self.config.SHEET_NAMES['improvement'])
        
        # Headers
        ws.append(self.config.HEADERS['improvement'])
        
        # Format headers
        for cell in ws[1]:
            cell.font = self.config.HEADER_FONT
            cell.fill = self.config.HEADER_FILL
            cell.alignment = self.config.CENTER_ALIGNMENT
        
        # Add improvement plan data
        for pillar_name, questions in pillars.items():
            for question in questions:
                if question.get('improvement_items'):
                    for item_data in question['improvement_items']:
                        ws.append([
                            question['pillar'],
                            question['question_id'],
                            question['question'],
                            question['risk_level'],
                            item_data['item'],
                            item_data['url']
                        ])
                        
                        # Color code risk level
                        current_row = ws.max_row
                        risk_cell = ws.cell(row=current_row, column=4)
                        if question['risk_level'] in self.config.RISK_COLORS:
                            risk_cell.fill = self.config.RISK_COLORS[question['risk_level']]
                        
                        # Make URL clickable
                        if item_data['url']:
                            url_cell = ws.cell(row=current_row, column=6)
                            url_cell.hyperlink = item_data['url']
                            url_cell.style = "Hyperlink"
        
        # Set column widths
        widths = self.config.COLUMN_WIDTHS['improvement']
        ws.column_dimensions['A'].width = widths['A']
        ws.column_dimensions['B'].width = widths['B']
        ws.column_dimensions['C'].width = widths['C']
        ws.column_dimensions['D'].width = widths['D']
        ws.column_dimensions['E'].width = widths['E']
        ws.column_dimensions['F'].width = widths['F']
    
    def _create_pillar_sheets(self, wb, pillars):
        """Create individual pillar sheets"""
        for pillar_name, questions in pillars.items():
            if questions:
                ws = wb.create_sheet(pillar_name)
                
                # Headers
                ws.append(self.config.HEADERS['pillar'])
                
                # Format headers
                for cell in ws[1]:
                    cell.font = self.config.HEADER_FONT
                    cell.fill = self.config.HEADER_FILL
                    cell.alignment = self.config.CENTER_ALIGNMENT
                
                current_row = 2
                
                for question in questions:
                    # Calculate choice statistics
                    selected_count = question['stats']['selected']
                    not_selected_count = question['stats']['not_selected']
                    na_count = question['stats']['na']
                    choice_stats = f"✅ {selected_count} | ⚠️ {not_selected_count} | ❌ {na_count}"
                    
                    # Main question row
                    ws.append([
                        question['question_id'],
                        question['question'],
                        choice_stats,
                        question['risk_level'],
                        question['notes'],
                        question.get('improvement_plan', '')
                    ])
                    
                    # Format question row
                    for col in range(1, 7):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = self.config.QUESTION_FONT
                        cell.fill = self.config.QUESTION_FILL
                        
                        if col == 3:  # Center choice/details column
                            cell.alignment = self.config.CENTER_ALIGNMENT
                    
                    # Color code risk level
                    if question['risk_level']:
                        risk_cell = ws.cell(row=current_row, column=4)
                        if question['risk_level'] in self.config.RISK_COLORS:
                            risk_cell.fill = self.config.RISK_COLORS[question['risk_level']]
                    
                    current_row += 1
                    group_start = current_row
                    
                    # Add choices as sub-rows
                    for choice in question['choices']:
                        ws.append([
                            '',
                            f"   • {choice['choice']}",
                            choice['status'],
                            '',
                            choice.get('description', ''),
                            ''
                        ])
                        
                        # Format choice row
                        status_cell = ws.cell(row=current_row, column=3)
                        status_cell.alignment = self.config.CENTER_ALIGNMENT
                        
                        if choice['status'] in self.config.STATUS_COLORS:
                            status_cell.fill = self.config.STATUS_COLORS[choice['status']]
                        
                        current_row += 1
                    
                    # Group choices under question
                    if len(question['choices']) > 0:
                        group_end = current_row - 1
                        try:
                            ws.row_dimensions.group(group_start, group_end, outline_level=1, hidden=True)
                        except:
                            pass
                
                # Set column widths
                widths = self.config.COLUMN_WIDTHS['pillar']
                ws.column_dimensions['A'].width = widths['A']
                ws.column_dimensions['B'].width = widths['B']
                ws.column_dimensions['C'].width = widths['C']
                ws.column_dimensions['D'].width = widths['D']
                ws.column_dimensions['E'].width = widths['E']
                ws.column_dimensions['F'].width = widths['F']
                
                # Set row heights
                for row in ws.iter_rows():
                    ws.row_dimensions[row[0].row].height = 15
